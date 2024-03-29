import csv
import sys
from collections import OrderedDict
from datetime import datetime
from decimal import Decimal, getcontext
from os.path import join, dirname, isfile, splitext
from os import makedirs

from dateutil.parser import parser
from openpyxl import load_workbook

from ExcelUtils import find_value, addlist

PARSER = parser()


def transplusminus(invalue):
    if invalue.lower() == 'af':
        return '-'
    return '+'


def setamount(inamount, decnr:int = 0):
    if ',' in inamount:
        inamount = inamount.replace(',', '.')
    elif '.' not in inamount and ',' not in inamount:
        decnr *= -1
        inamount = inamount[:decnr] + '.' + inamount[decnr:]
    return Decimal(inamount)


ING_mapping = [(0, 'transactiondate', lambda x: PARSER.parse(x)),
               (1, 'naam', None),
               (2, 'ownaccountnr', None),
               (3, 'accountnr', None),
               (5, 'plusminus', lambda x: transplusminus(x)),
               (6, 'amount', lambda x: setamount(x,2)),
               ]


class ImportBase:
    transactiondate = datetime.now()
    parent = None
    verwerkingsjaar = None
    plusminus = '+'
    amount = Decimal('0')
    accountnr = '0'
    naam = 'Onbekend'
    ownaccountnr = None
    grootboekrekening = None

    def __init__(self, administratie: str, bank_sjabloon: list = None, verwerkingsjaar: int = None,
                 vorigjaar: str = None, importfile: str = None):
        self.rowlist = []
        self.vorigjaar = vorigjaar
        if vorigjaar:
            self.wb_vorigjaar = load_workbook(self.vorigjaar, keep_vba=True)

        if not verwerkingsjaar:
            nam, ext = splitext(administratie)
            self.verwerkingsjaar = int(nam[-4:])
        else:
            self.verwerkingsjaar = verwerkingsjaar
        self.administratie = administratie
        self.import_all = False

        if not isfile(administratie):
            administratie = join(dirname(dirname(__file__)), 'resources', 'Administratie_sjabl.xlsx')
            self.import_all = True
        self.wb = load_workbook(administratie)
        self.rekeningschema = self.buildrekschema(self.wb['Rekeningschema'])
        common_ws = self.wb['Standaardwaarden']
        self.contributiebedrag = self.bldcontributie(common_ws)

        self.importfile = importfile
        self.bank_sjabloon = bank_sjabloon

    def bldcontributie(self, common_ws):
        result = OrderedDict()
        cell = find_value('Contributiebedrag', common_ws)
        datekey = common_ws['C' + str(cell.row)].value
        if datekey:
            result[datekey] = Decimal(common_ws['D' + str(cell.row)].value)
            result[None] = Decimal(common_ws['B' + str(cell.row)].value)
        else:
            result[None] = Decimal(common_ws['B' + str(cell.row)].value)
        return result

    def contributie_per_maand(self, maand: int):
        fc = [c for d, c in self.contributiebedrag.items() if d and d > maand or not d]
        if len(fc):
            cbedrag = fc[0]
        else:
            cbedrag = list(self.contributiebedrag.values())[0]
        return cbedrag

    def process_transactions(self):
        transws = self.wb['Transacties']
        previous_value = None
        rowcount = 0
        for newrow in self.rowlist:
            accountcell = None
            if not self.import_all:
                accountcell = find_value([c for c in [newrow.accountnr, newrow.amount, newrow.transactiondate] if c], transws)
            if not accountcell:
                if previous_value and previous_value != newrow.ownaccount:
                    blankrow = 1
                else:
                    blankrow = 0
                rowcount += 1
                addlist(transws, newrow.addtransrow(), blankrow=blankrow)
        print('%s rijen toegevoegd in transacties' % rowcount)
        return transws

    def process_leden(self):
        ledenws = self.wb['Leden']
        for newrow in [r for r in self.rowlist if r.is_contributie()]:
            accountcell = find_value(newrow.accountnr, ledenws, max_col=1)
            if not accountcell:
                addlist(ledenws, newrow.addledenrow())
        return ledenws

    def buildrekschema(self, rekws):
        rekeningschema = {}
        for rekeningschemarij in rekws.rows:
            v = rekeningschemarij[0].value
            if v:
                try:
                    rekeningnr = int(v)
                except ValueError:
                    continue
                rekeningschema[rekeningnr] = {
                    'omschrijving': rekeningschemarij[2].value,
                    'herkenningswaarden': {
                        'plusminus': '+' if len(rekeningschemarij) > 3 and rekeningschemarij[3].value and
                                            rekeningschemarij[3].value.lower() == 'bij' else '-',
                        'naam': rekeningschemarij[4].value.lower() if len(rekeningschemarij) > 4 and rekeningschemarij[
                            4].value else None,
                        'min_bedrag': Decimal(rekeningschemarij[5].value) if len(rekeningschemarij) > 5 and
                                                                             rekeningschemarij[5].value else None,
                        'max_bedrag': Decimal(rekeningschemarij[6].value) if len(rekeningschemarij) > 6 and
                                                                             rekeningschemarij[6].value else None}}
        return rekeningschema

    def getkey(self):
        return str(self.ownaccountnr) + self.transactiondate.strftime('%Y%m%d')

    def __repr__(self):
        return str(self.ownaccountnr)

    def checkowngrbrek(self):
        ok_recognized = {}
        for grootboekrekening, grbkwaarde in self.parent.rekeningschema.items():
            ok_recognized[grootboekrekening] = [0, 0, 0, 0]
            idx = -1
            minwaarde = Decimal('0')
            zwaarte = 100
            for herkenningsitem, waarde in grbkwaarde['herkenningswaarden'].items():
                idx += 1
                if not waarde:
                    continue
                if isinstance(waarde, str):
                    listwaarde = [w.strip().lower() for w in waarde.split(',')]
                else:
                    listwaarde = [waarde]
                if herkenningsitem in self.__dict__:
                    for lw in listwaarde:
                        currentval = getattr(self, herkenningsitem)
                        if lw == '*' or (lw and currentval and lw in currentval.lower()):
                            ok_recognized[grootboekrekening][idx] = zwaarte
                    zwaarte = int(zwaarte / 2)
                elif idx == 2 and waarde <= self.amount:
                    minwaarde = waarde
                    ok_recognized[grootboekrekening][idx] = zwaarte
                elif idx == 3 and waarde >= self.amount >= minwaarde:
                    ok_recognized[grootboekrekening][idx] = zwaarte
        found_grootboekrekening = None
        max_waarde = 0
        for grootboekrekening, found_values in ok_recognized.items():
            if found_values[0] and found_values[1] and max_waarde < sum(found_values):
                max_waarde = sum(found_values)
                found_grootboekrekening = grootboekrekening
        if max_waarde > 120:
            return found_grootboekrekening

    def is_contributie(self):
        cbedrag = self.parent.contributie_per_maand(self.transactiondate.month)
        equal = True
        if cbedrag > self.amount:
            equal = cbedrag % self.amount < 2
        elif cbedrag < self.amount:
            equal = self.amount % cbedrag < 2
        return self.plusminus == '+' and self.amount and equal

    def addledenrow(self):
        return [self.accountnr, self.naam, "", "", "", self.transactiondate, ""]

    def addtransrow(self):
        return [self.transactiondate, self.ownaccountnr, self.grootboekrekening, self.accountnr, self.plusminus,
                self.amount, self.naam]

    def save(self):
        if not isfile(self.administratie):
            makedirs(dirname(self.administratie), exist_ok=True)
        self.wb.save(self.administratie)

    def bouw_vanuit_vorigjaar(self):
        self.vorigjaar = load_workbook(self.vorigjaar, data_only=True)
        for idx, row in enumerate(self.vorigjaar['Resultaten']):
            if idx > 3:
                if self.vorigjaar['Resultaten'][idx+1][1].value and \
                        self.vorigjaar['Resultaten'][idx+1][1].value.lower() == 'resultaat':
                    break
                resultcell = find_value(self.vorigjaar['Resultaten'][idx+1][0].value, self.wb['Resultaten'])
                if resultcell:
                    self.wb['Resultaten'][resultcell.row][2].value = self.vorigjaar['Resultaten'][resultcell.row][3].value
                else:
                    addlist(self.wb['Resultaten'], [row[0].value, row[1].value, row[3].value])

        for idx, row in enumerate([lid for lid in self.vorigjaar['Leden'] if lid[7].value is None]):
            addlist(self.wb['Leden'], [row[0].value, row[1].value, row[2].value, row[3].value, row[4].value])


class ImportLineXLS(ImportBase):
    def __init__(self, parent, inlist):
        self.transactiondate = datetime.now()
        self.parent = parent
        if self.parent.verwerkingsjaar:
            self.verwerkingsjaar = self.parent.verwerkingsjaar
        else:
            self.verwerkingsjaar = datetime.now().year
        self.plusminus = '+'
        self.amount = Decimal('0')
        self.accountnr = '0'
        self.naam = 'Onbekend'
        self.ownaccountnr = None
        if not parent.bank_sjabloon:
            parent.bank_sjabloon = ING_mapping
        for colnr, fieldname, conversion in parent.bank_sjabloon:
            setattr(self, fieldname, conversion(inlist[colnr]) if conversion else inlist[colnr])
        if self.transactiondate.year != self.verwerkingsjaar:
            sys.exit('Er zijn transacties van een ander jaar aangetroffen')
        self.grootboekrekening = self.checkowngrbrek()
        self.amount = self.amount if self.plusminus == '+' else self.amount * -1


class ImportLineCSV(ImportBase):
    def __init__(self, parent, inlist):
        self.transactiondate = datetime.now()
        self.parent = parent
        if self.parent.verwerkingsjaar:
            self.verwerkingsjaar = self.parent.verwerkingsjaar
        else:
            self.verwerkingsjaar = datetime.now().year
        self.plusminus = '+'
        self.amount = Decimal('0')
        self.accountnr = '0'
        self.naam = 'Onbekend'
        self.ownaccountnr = None
        if not parent.bank_sjabloon:
            parent.bank_sjabloon = ING_mapping
        for colnr, fieldname, conversion in parent.bank_sjabloon:
            setattr(self, fieldname, conversion(inlist[colnr]) if conversion else inlist[colnr])
        if self.transactiondate.year != self.verwerkingsjaar:
            sys.exit('Er zijn transacties van een ander jaar aangetroffen')
        self.grootboekrekening = self.checkowngrbrek()
        self.amount = self.amount if self.plusminus == '+' else self.amount * -1

# class ImportBase:
#     def __init__(self, importfile, administratie: str, bank_sjabloon: list = None, verwerkingsjaar: int = None,
#                  vorigjaar: str = None):
#         self.rowlist = []
#         self.vorigjaar = vorigjaar
#         if vorigjaar:
#             self.wb_vorigjaar = load_workbook(self.vorigjaar, keep_vba=True)
#         if not verwerkingsjaar:
#
#         self.verwerkingsjaar = verwerkingsjaar
#         self.administratie = administratie
#         self.import_all = False
#
#         if not isfile(administratie):
#             administratie = join(dirname(dirname(__file__)), 'resources', 'Administratie_sjabl.xlsx')
#             self.import_all = True
#         self.wb = load_workbook(administratie)
#         common_ws = self.wb['Standaardwaarden']
#         self.contributiebedrag = self.bldcontributie(common_ws)
#         self.rekeningschema = self.buildrekschema(self.wb['Rekeningschema'])
#
#         self.importfile = importfile
#         self.bank_sjabloon = bank_sjabloon
#
#     def calc_contributie_per_lid(self):
#         ledenws = self.wb['Leden']
#         save_functie = None
#         for rijnr, lidrij in enumerate(ledenws.rows):
#             if lidrij[8].value:
#                 save_functie = lidrij[8].value
#             contributie = Decimal('0')
#             rekeningnr = lidrij[0]
#             lid_sinds = lidrij[4]
#             lid_tot = lidrij[6]
#             if rekeningnr:
#                 if lid_sinds.value:
#                     try:
#                         maand = lid_sinds.value.month
#                     except AttributeError:
#                         continue
#                 else:
#                     maand = 1
#                 if lid_tot.value:
#                     maand_tot = lid_tot.value.month
#                 else:
#                     maand_tot = datetime.now().month + 1
#                 while maand < maand_tot:
#                     contributie += Decimal(self.contributie_per_maand(maand))
#                     maand += 1
#                 list(lidrij)[7].value = contributie
#                 if save_functie and rijnr != 2:
#                     repa = 'A' + str(rijnr + 1)
#                     reph = 'H' + str(rijnr + 1)
#                     list(lidrij)[8].value = save_functie.replace('A3', repa).replace('H3', reph)
#         return ledenws
#
#     def calc_openstaande_contributie(self):
#         ledenws = self.wb['Leden']
#         transws = self.wb['Transacties']
#         betaald = Decimal('0')
#         for lid in ledenws.rows:
#             if lid[0].value and isinstance(lid[0].value, int):
#                 for trans in transws.rows:
#                     if trans[2] == lid[0] and trans[3] == '+' and self.contributie_per_maand(trans[0].month) % trans[
#                         4] < 2:
#                         betaald += trans[4]
#                 list(lid)[8].value = lid[7].value - betaald
#         return ledenws
#
#     def proc_debiteuren(self):
#         transws = self.wb['Transacties']
#         debws = self.wb['Debiteuren']
#         for deb in debws.rows:
#             try:
#                 openstaand_bedrag = Decimal(deb[2].value)
#             except:
#                 continue
#             if deb[1].value:
#                 naam = deb[1].value.lower()
#                 cell = find_value(openstaand_bedrag, transws)
#
#                 if cell and transws[cell.rownr][1].value and naam in transws[cell.rownr][1].value.lower():
#                     list(deb)[4].value = transws[cell.rownr][0]
#                     transws[cell.rownr][6].value = deb[0].value
#         return debws
#

class ImportCsv(ImportBase):
    def __init__(self, importfile, administratie: str, bank_sjabloon: list = None, verwerkingsjaar: int = None,
                 vorigjaar: str = None):
        super().__init__(administratie, bank_sjabloon, verwerkingsjaar, vorigjaar, importfile)

    def process_import_lines(self, delim):
        error = False
        with open(self.importfile) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=delim)
            for line_count, row in enumerate(csv_reader):
                if line_count == 0:  # HEADER
                    print('Import van csv bestand %s loopt' % csv_file)
                else:  # DATALINES
                    try:
                        row_object = ImportLineCSV(self, row)
                    except Exception:
                        error = True
                        break
                    # if self.checkdouble(row_object):
                    #     continue
                    if not self.verwerkingsjaar:
                        self.verwerkingsjaar = row_object.verwerkingsjaar
                    self.rowlist.append(row_object)
        return error

    def process_importfile(self):
        delimiters = (',', ';', ':')
        for delim in delimiters:
            error = self.process_import_lines(delim)
            if not error:
                break
        else:
            sys.exit('Juiste delimiter voor csv kon niet worden gevonden, verwerking is onmogelijk')

        self.rowlist.sort(key=lambda x: x.getkey())
        self.process_transactions()
